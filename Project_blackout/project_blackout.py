"""
Takes in an excel file with group shift stats on a particular month
Takes as input another excel file as a request graph
For each provider matching provider, blacksout each date onto another excel file along with an indicator text marking 'w' for work concatenated with extra text if shift ends at 7pm or later
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime


def main():

    inp1 = input(f"Please input location path of group shift stats: ")
    PATH_group_stats = fr"{return_path(inp1)}"

    inp2 = input(f"Please input location path of request graph: ")
    PATH_req_graph = fr"{return_path(inp2)}"

    df1 = return_df(PATH_group_stats)
    df2 = return_df(PATH_req_graph)

    df = process_shift_data(df1)

    df_r = process_request_data(df, df2)

    fill_cells(PATH_req_graph, df_r)

    print(f"SUCCESS! Saved as : group_assigned_shifts_and_requests.xlsx ")


def return_path(x):
    return os.path.abspath(x.strip()[1:-1])


def return_df(p):
    return pd.read_excel(p, header=1)


def load_sheets(PATH, f, s):
    sheet_name, file_name = (s, f)
    wb = load_workbook(PATH, data_only=True)
    wb.save(filename=f"{file_name}")
    return wb[f"{sheet_name}"], wb


def process_shift_data(df1):
    # reads the shift excel file given as an argument and creates a dataframe

    # reads excel with header
    df = df1

    # adds f_name, l_name, & end_time
    name_data = df["Provider"].str.split(" ", n=1, expand=True)
    df["f_name"] = name_data[0]
    df["l_name"] = name_data[1]
    time_data = df["Time"].str.split(" - ", n=1, expand=True)
    df["end_time"] = time_data[1] + "m"

    # changes dtype obj->datetime
    df["Date"] = pd.to_datetime(df["Date"])
    df["end_time"] = pd.to_datetime(df["end_time"], format="%I%p")

    # drop rows/Providers without shifts where df.Shift = NaN
    df = df[df["Shift"].notna()]
    df = df.reset_index(drop=True)

    return df


def process_request_data(df, df2):
    df = df
    df_r = df2

    # delete On / off / time of columns and then reset index.
    # ONLY has providers with scheduled shifts
    df_r = df_r.drop(df_r.tail(4).index)
    df_r.drop(columns=["On", "Off", "Time Of"])

    # for each name in df, add shifts from df to df_request (df_r)
    for _, row in df.iterrows():
        name = row["Provider"]
        day = row["Date"].day
        end_time = row["end_time"]

        # TODO create a datetime object to simplify below

        late_times = [
            "20:00:00",
            "21:00:00",
            "22:00:00",
            "23:00:00",
            "00:00:00",
            "01:00:00",
            "02:00:00",
            "06:00:00",
            "07:00:00",
        ]

        # if provider is in df_r, otherwise add
        if name not in df_r.User.tolist():

            # If Provider in df doesn't exist in df_r, create row in df with 'User':name
            new_row = pd.DataFrame({"User": name}, index=[-1])
            df_r2 = pd.concat([df_r.loc[:], new_row]).reset_index(drop=True)
            df_r = df_r2.copy()

        # find index of provider in df
        provider_index = df_r.User[df_r.User == name].index.tolist()[0]

        # write onto the cell with an indicator
        indicator = "w"
        if str(end_time.time()) in late_times:
            indicator += "til" + str(end_time.time())[:5]
        df_r.at[provider_index, str(day)] = indicator

    return df_r


# TODO add prior month day columns ~3 days +/- shade
def get_dict_hex_data(sh):
    # get  dict(hex_data) for each provider for each cell
    provider_hex_d = dict(
        zip(
            [
                row[0]
                for row in sh.iter_rows(
                    min_row=3, max_row=(sh.max_row - 4), max_col=(1), values_only=True
                )
            ],
            [
                [cell.fill.start_color.index for cell in row]
                for row in sh.iter_rows(
                    min_col=2,
                    max_col=(sh.max_column - 4),
                    min_row=3,
                    max_row=(sh.max_row - 4),
                )
            ],
        )
    )
    return provider_hex_d

def fill_cells(PATH_req_graph, df_r):
    sh, _ = load_sheets(PATH_req_graph, "temp.xslx", "Request Graph")
    df_r.to_excel("pandas_processed1.xlsx")
    sh_r, wb_r = load_sheets("pandas_processed1.xlsx", "temp2.xslx", "Sheet1")
    provider_hex_d = get_dict_hex_data(sh)
    color_d = {
        "10": "00C0C0C0",
        "11": "800080", # 11 is specific time off dates
        "12": "00b200",
        "9": "ff8b94", #9 is blackout dates
        "00000000": "00FFFFFF",
    }

    for row in sh_r.iter_rows(
        min_col=2, max_col=(sh_r.max_column - 1), min_row=2, max_row=(sh_r.max_row)
    ):
        key = row[0].value

        try:

            for i in range(len(provider_hex_d[key])):
                # fill cell  and move to next cell
                cellc_to_fill = row[i + 1].coordinate

                code = str(provider_hex_d[key][i])
                sh_r[cellc_to_fill].fill = PatternFill(
                    fgColor=color_d[code], fill_type="solid"
                )

        except KeyError:
            pass

    now = datetime.now()
    s = f"{now.month:02d}_{now.day}_{now.hour}{now.minute}"
    PATH_export = os.path.abspath(
        fr"C:\Users\Angel\Downloads\{s}group_assigned_shifts_and_requests.xlsx"
    )
    wb_r.save(filename=PATH_export)

    os.remove("temp.xslx")
    os.remove("temp2.xslx")
    os.remove("pandas_processed1.xlsx")


if __name__ == "__main__":
    main()