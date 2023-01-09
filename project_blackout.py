"""
Reads excel file
For each provider, blacks out each date onto another excel file
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import sys
import os
from datetime import datetime

def main():

    inp1 = input(f"Please input location path of group shift stats: ").strip()
    PATH_group_stats = fr'{inp1[1:-1]}'

    inp2 = input(f"Please input location path of request graph: ").strip()
    PATH_req_graph = fr'{inp2[1:-1]}'

    PATH_request_processed1 = os.path.abspath(r"C:\Users\Angel\anaconda3\envs\pandas_processed1.xlsx")


    process_shift_data(PATH_group_stats, PATH_req_graph)
    fill_cells(PATH_req_graph, PATH_request_processed1)

    print(f"SUCCESS! Saved as : group_assigned_shifts_and_requests.xlsx ")

def process_shift_data(PATH_group_stats, PATH_req_graph):
    # reads the shift excel file given as an argument and creates a dataframe

    # reads excel with header
    df = pd.read_excel(PATH_group_stats, header =1)

    #adds f_name, l_name, & end_time
    name_data = df['Provider'].str.split(" ", n=1, expand=True)
    df['f_name'] = name_data[0]
    df['l_name'] = name_data[1]
    time_data = df['Time'].str.split(" - ", n=1, expand=True)
    df['end_time'] = time_data[1] + 'm'

    #changes dtype obj->datetime
    df['Date'] = pd.to_datetime(df['Date'])
    df['end_time'] = pd.to_datetime(df['end_time'], format='%I%p')

    # drop rows/Providers without shifts where df.Shift = NaN
    df = df[df['Shift'].notna()]
    df = df.reset_index(drop=True)

    process_request_data(df, PATH_req_graph)


def process_request_data(df, PATH_req_graph):
    
    df_r = pd.read_excel(PATH_req_graph, header =1)

    # delete On / off / time of columns and then reset index.
    # ONLY has providers with scheduled shifts  
    df_r = df_r.drop(df_r.tail(4).index)
    df_r.drop(columns=['On','Off','Time Of'])

    # for each name in df, add shifts from df to df_request (df_r)
    for i, row in df.iterrows():
        name = row['Provider']
        day = row['Date'].day
        end_time = row['end_time']
        late_times = ['20:00:00', '21:00:00', '22:00:00', '23:00:00', '01:00:00', '02:00:00', '06:00:00']

        
        # if provider is in df_r, otherwise add
        if name not in df_r.User.tolist():

            # If Provider in df doesn't exist in df_r, create row in df with 'User':name
            new_row = pd.DataFrame({'User': name}, index=[-1])
            df_r2 = pd.concat([df_r.loc[:], new_row]).reset_index(drop=True)
            df_r = df_r2.copy()

        # find index of provider in df
        provider_index = df_r.User[df_r.User == name ].index.tolist()[0]

        # write onto the cell with an indicator
        indicator = 'w'
        if str(end_time.time()) in late_times:
            indicator += ' til' + str(end_time.time())[:5]
        df_r.at[provider_index, str(day)] = indicator

    # save excel file 
    df_r.to_excel('pandas_processed1.xlsx')


# TODO add prior month day columns ~3 days +/- shade 

# TODO For each provider, for each shift, color code cells Black or next cell Red if df.end_time is 8p or after.

def fill_cells(PATH_req_graph, PATH_request_processed1):
    wba = load_workbook(PATH_req_graph, data_only = True)
    wba.save(filename='temp.xslx')
    sh = wba['Request Graph']


    wb_r = load_workbook(PATH_request_processed1, data_only=True)
    wb_r.save(filename = 'temp2.xslx')
    sh_r = wb_r['Sheet1']

    # get  dict(hex_data) for each provider for each cell
    provider_hex_d = dict(zip(
        [row[0] for row in sh.iter_rows(min_row=3, max_row=(sh.max_row - 4), max_col=(1), values_only=True)], 

        [[cell.fill.start_color.index  for cell in row] for row in sh.iter_rows(min_col=2, max_col=(sh.max_column-4), min_row=3, max_row=(sh.max_row - 4))]
    )) 

    color_d = {'10': '00C0C0C0' ,'11':'00808080', '9': 'ff8b94', '00000000': '00FFFFFF' }

    for row in sh_r.iter_rows(min_col=2, max_col=(sh_r.max_column - 1), min_row=2, max_row=(sh_r.max_row)):
        key = row[0].value

        try:

            for i in range(len(provider_hex_d[key])):
                #fill cell  and move to next cell
                cellc_to_fill = row[i + 1].coordinate

                code = str(provider_hex_d[key][i])
                sh_r[cellc_to_fill].fill = PatternFill(fgColor=color_d[code], fill_type = "solid")

        except KeyError:
            pass
    
    PATH_export = os.path.abspath(r'C:\Users\Angel\Downloads\group_assigned_shifts_and_requests.xlsx')
    wb_r.save(filename = PATH_export)

if __name__ == "__main__":
    main()